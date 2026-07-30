# ruff: noqa: E501, F821, F841, N803, N806

import marimo

__generated_with = "0.20.4"
app = marimo.App(width="medium")


@app.cell(hide_code=True)
def _():
    import ast
    import importlib.util
    import os
    from pathlib import Path
    from typing import Literal

    import marimo as mo
    from dotenv import load_dotenv
    from openpyxl import load_workbook
    from predict_rlm import File, PredictRLM
    from pydantic import BaseModel, Field
    from skills import evidence_coding_skill

    import avalanche as ava

    return (
        BaseModel,
        Field,
        File,
        Literal,
        Path,
        PredictRLM,
        ast,
        ava,
        evidence_coding_skill,
        importlib,
        load_dotenv,
        load_workbook,
        mo,
        os,
    )


@app.cell(hide_code=True)
def _(Path, ast, importlib, load_dotenv, os):
    load_dotenv()

    MODEL = os.getenv("CUSTOMER_FEEDBACK_REVIEW_MODEL", "openai/gpt-5.6-terra")
    SUB_MODEL = os.getenv(
        "CUSTOMER_FEEDBACK_REVIEW_SUB_MODEL", "gemini/gemini-3.5-flash"
    )
    FEEDBACK_WORKBOOK_PATH = Path(__file__).with_name("feedback_workbook.xlsx")
    WORKFLOW_DAG_PATH = Path(__file__).with_name("workflowdag.jpg")
    WORKFLOW_NODE_TYPES_DAG_PATH = Path(__file__).with_name("workflowdag2.jpg")
    ARTIFACT_ROOT = Path(__file__).with_name("artifacts") / "generated_review_pack"
    WORKBOOK_OUTPUT_DIR = ARTIFACT_ROOT / "workbook"
    BRIEF_OUTPUT_DIR = ARTIFACT_ROOT / "brief"

    def _source_symbols(source: str, *symbol_names: str) -> str:
        source_lines = source.splitlines()
        tree = ast.parse(source)
        nodes = {}
        for node in tree.body:
            if isinstance(node, (ast.ClassDef, ast.FunctionDef, ast.AsyncFunctionDef)):
                nodes[node.name] = node
            elif isinstance(node, ast.Assign):
                for target in node.targets:
                    if isinstance(target, ast.Name):
                        nodes[target.id] = node
        blocks = []
        for symbol_name in symbol_names:
            node = nodes[symbol_name]
            decorators = (
                node.decorator_list
                if isinstance(node, (ast.ClassDef, ast.FunctionDef, ast.AsyncFunctionDef))
                else []
            )
            first_line = min([node.lineno, *(item.lineno for item in decorators)])
            blocks.append("\n".join(source_lines[first_line - 1 : node.end_lineno]))
        return "\n\n\n".join(blocks)

    def source_code(file_name: str, *symbol_names: str) -> str:
        source = Path(__file__).with_name(file_name).read_text(encoding="utf-8")
        return _source_symbols(source, *symbol_names)

    def module_source_code(module_name: str, *symbol_names: str) -> str:
        spec = importlib.util.find_spec(module_name)
        if spec is None or spec.origin is None:
            raise RuntimeError(f"Cannot locate source module {module_name!r}")
        source = Path(spec.origin).read_text(encoding="utf-8")
        return _source_symbols(source, *symbol_names)

    return (
        BRIEF_OUTPUT_DIR,
        FEEDBACK_WORKBOOK_PATH,
        MODEL,
        SUB_MODEL,
        WORKBOOK_OUTPUT_DIR,
        WORKFLOW_DAG_PATH,
        WORKFLOW_NODE_TYPES_DAG_PATH,
        module_source_code,
        source_code,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.vstack(
        (
            mo.md("# From customer feedback to a product review pack"),
            mo.md(
                "A product team receives customer feedback faster than it can review it. "
                "Important requests repeat across accounts, retention risks are buried in "
                "individual comments, and the supporting account and roadmap context lives "
                "in separate workbook tabs."
            ),
            mo.callout(
                "Given a bundled customer-feedback workbook, produce recurring themes, "
                "material account risks, a review workbook, and an executive brief—with "
                "every conclusion traceable to its source feedback.",
                kind="info",
            ),
        )
    )
    return


@app.cell(hide_code=True)
def _(FEEDBACK_WORKBOOK_PATH, load_workbook, mo):
    _workbook = load_workbook(FEEDBACK_WORKBOOK_PATH, read_only=True, data_only=False)
    _feedback_sheet = _workbook["Feedback"]
    _headers = [cell.value for cell in _feedback_sheet[1]]
    _feedback_rows = [
        dict(zip(_headers, row, strict=True))
        for row in _feedback_sheet.iter_rows(min_row=2, values_only=True)
    ]
    _workbook.close()

    mo.vstack(
        (
            mo.md("## 1. Start with the real input"),
            mo.md(
                "The example ships with this workbook. Its `Feedback` sheet appears below; "
                "the same file also contains `Accounts` and `Roadmap` sheets. At this point "
                "there are no nodes, decorators, or agents—only the source material and the "
                "outcome we need."
            ),
            mo.ui.table(_feedback_rows),
        )
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.vstack(
        (
            mo.md("## 2. Describe the work before the implementation"),
            mo.md(
                "A useful workflow starts as a set of logical responsibilities. Each one "
                "has a result we can name and inspect; none is a framework primitive yet."
            ),
            mo.md(
                """
                1. **Load the feedback workbook** as one file.
                2. **Find recurring product themes** with traceable evidence.
                3. **Identify material account risks** using feedback and account context.
                4. **Combine both reports** into one typed product review.
                5. **Build the review workbook** from the approved review.
                6. **Write the executive brief** from the same approved review.
                7. **Publish both artifacts** as one review pack.
                """
            ),
        )
    )
    return


@app.cell(hide_code=True)
def simple_live_rlm(WORKFLOW_DAG_PATH, mo):
    mo.vstack(
        (
            mo.md("### Dependencies come from the data"),
            mo.md(
                "Theme extraction and risk detection both consume the same workbook. The "
                "approved review then feeds two independent rendering tasks before their "
                "artifacts are published together."
            ),
            mo.image(
                WORKFLOW_DAG_PATH,
                alt="Customer feedback product review workflow DAG",
                width="100%",
                rounded=True,
            ),
        )
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.vstack(
        (
            mo.md("## 3. Zoom in: find recurring product themes"),
            mo.md(
                "The agent must inspect the workbook, compare comments across accounts, "
                "decide which needs genuinely recur, calculate counts from source rows, "
                "and preserve the workbook evidence behind every conclusion."
            ),
            mo.callout(
                "This is one logical agent task. Inspection, classification, aggregation, "
                "and validation are parts of its internal strategy—not separate workflow "
                "nodes.",
                kind="warn",
            ),
        )
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""
    ### Define the task

    **Contract**: What should the agent's contract be? What does it receive, what should it do with it and how, and what should it return?

    | Boundary | Contract |
    |---|---|
    | **Input** | The complete feedback workbook as a `File` |
    | **Responsibility** | Find recurring product needs and cite their workbook evidence |
    | **Strategy** | Inspect, classify, aggregate, cross-check, and validate |
    | **Output** | A validated `ThemeReport` |

    **Capabilities**: What knowledge & capabilities does it need? What services does it need to reach?

    | Surface | Capability |
    |---|---|
    | **Knowledge** | Spreadsheet handling |
    | **Services** | None |
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""
    ## PredictRLM: Avalanche's agent runtime

    Avalanche runs each agent step with PredictRLM. It gives the outer model a
    stateful Python environment: on each turn, the model's only action is to
    write Python, observe what that code returns, and continue until it submits
    the typed result.

    **Organizing context.** Inputs, files, and intermediate results live in
    Python variables. The agent chooses what to inspect or pass into a focused
    subcall, so its entire working set is not dumped into every model context.

    **Organizing tool calls.** Tools are ordinary Python functions. One model
    turn can emit code that calls several tools, loops over their results,
    combines them, and preserves useful state for the next turn.

    ### Example PredictRLM turn

    ```python
    RLM turn 1/30 (ok)
      reasoning: I need to inspect the workbook before semantic classification. I should
                 locate the mounted file, list worksheets, validate headers, and count rows.
      python: 8 lines
      output: 214 chars
      code:
        from pathlib import Path
        import pandas as pd
        workbook_path = next(Path("/sandbox/input/workbook").glob("*.xlsx"))
        excel = pd.ExcelFile(workbook_path)
        print("sheets", excel.sheet_names)
        feedback = pd.read_excel(workbook_path, sheet_name="Feedback", dtype={"feedback_id": str, "account_id": str})
        print("rows", len(feedback), "columns", feedback.columns.tolist())
        print(feedback.head(3).to_dict(orient="records"))
    ```
    """)
    return


@app.cell(hide_code=True)
def _(mo, source_code):
    mo.vstack(
        (
            mo.md("## 4. The anatomy of an `agent`"),
            mo.md("A PredictRLM instance is defined by **3 core components**:"),
            mo.md("### A. Signature: contract and strategy"),
            mo.md(
                "A signature defines the agent's **inputs**, **outputs** and "
                "**instructions**. Here the agent receives a workbook `File` and must "
                "return a `ThemeReport`."
            ),
            mo.ui.code_editor(
                value=source_code("signature.py", "ExtractThemes"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
            mo.ui.code_editor(
                value=source_code(
                    "schema.py",
                    "EvidenceReference",
                    "Theme",
                    "ThemeReport",
                ),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
        )
    )
    return


@app.cell(hide_code=True)
def _(mo, module_source_code, source_code):
    mo.vstack(
        (
            mo.md("### B. Skills: reusable sandbox knowledge"),
            mo.md(
                "The built-in spreadsheet skill supplies Excel mechanics, formulas, "
                "formatting, and validation. This custom skill adds a reusable procedure "
                "for turning qualitative records into evidence-backed findings; both "
                "analysis agents use it unchanged."
            ),
            mo.ui.code_editor(
                value=source_code("skills.py", "evidence_coding_skill"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
            mo.md("#### Built-in spreadsheet skill"),
            mo.ui.code_editor(
                value=module_source_code(
                    "predict_rlm.skills.spreadsheet.skill",
                    "spreadsheet_skill",
                ),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
        )
    )
    return


@app.cell(hide_code=True)
def tools_example(mo):
    mo.vstack(
        (
            mo.md("### C. Tools: typed host capabilities"),
            mo.md(
                "This workflow needs no host tool because account context is bundled in "
                "the workbook. If that context lived in an authenticated CRM, a typed host "
                "capability would expose it to PredictRLM as an ordinary Python function."
            ),
            mo.ui.code_editor(
                value='class AccountPlan(BaseModel):\n    account_id: str\n    tier: str\n\n\ndef lookup_account_plan(account_id: str) -> AccountPlan:\n    """Return the current CRM plan for one account."""\n    return crm_client.fetch_plan(account_id)',
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
        )
    )
    return


@app.cell(hide_code=True)
def quickstart_signature(mo, source_code):
    mo.vstack(
        (
            mo.md("## Putting it together: creating the agent"),
            mo.md(
                "`@ava.agent_step(...)` combines the signature with its model "
                "configuration and reusable skills. Avalanche injects the configured "
                "`ava.Agent`; the function passes it the file and validates the returned "
                "report."
            ),
            mo.ui.code_editor(
                value=source_code("flow.py", "extract_themes"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
        )
    )
    return


@app.cell(hide_code=True)
async def _(
    ExtractThemes,
    FEEDBACK_WORKBOOK_PATH,
    File,
    MODEL,
    PredictRLM,
    SUB_MODEL,
    ava,
    evidence_coding_skill,
    mo,
):
    _workbook = File(path=str(FEEDBACK_WORKBOOK_PATH))
    _rlm = PredictRLM(
        ExtractThemes,
        lm=MODEL,
        sub_lm=SUB_MODEL,
        skills=[ava.agent.skills.spreadsheet, evidence_coding_skill],
    )
    _result = await _rlm.acall(workbook=_workbook)

    mo.vstack(
        (
            mo.md("### Running that agent"),
            mo.inspect(_result.report.model_dump()),
        )
    )
    return


@app.cell(hide_code=True)
def quickstart_rlm_call(WORKFLOW_DAG_PATH, mo):
    mo.vstack(
        (
            mo.md("## 5. Zoom back out to the complete flow"),
            mo.md(
                "The theme agent is one branch, not the whole application. Put it back "
                "beside the other responsibilities and choose an execution type for each "
                "boundary."
            ),
            mo.image(
                WORKFLOW_DAG_PATH,
                alt="Customer feedback product review workflow DAG",
                width="100%",
                rounded=True,
            ),
        )
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.vstack(
        (
            mo.md("### Avalanche is the orchestration layer"),
            mo.md(
                "PredictRLM supplies the adaptive agent runtime. Avalanche places that "
                "runtime beside deterministic Python and external writes, then schedules "
                "the resulting graph from its declared data dependencies."
            ),
            mo.md("### Steps"),
            mo.md(
                "A normal typed Python function becomes a deterministic workflow node with "
                "`@ava.step`."
            ),
            mo.ui.code_editor(
                value="@ava.step\ndef normalize_feedback(text: str) -> str:\n    return text.strip()",
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
            mo.md("### Agent steps"),
            mo.md(
                "An `@ava.agent_step` wraps one logical agent task and receives its "
                "configured `ava.Agent` through dependency injection."
            ),
            mo.ui.code_editor(
                value="@ava.agent_step(AnalyzeWorkbook, skills=[ava.agent.skills.spreadsheet])\nasync def analyze_workbook(\n    workbook: File, *, agent: ava.Agent\n) -> WorkbookAnalysis:\n    return (await agent(workbook=workbook)).analysis",
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
            mo.md("### Workflows"),
            mo.md(
                "An `@ava.workflow` declares dependencies between nodes: `>>` orders work "
                "and `&` groups independent branches."
            ),
            mo.ui.code_editor(
                value="@ava.workflow\ndef example_workflow():\n    return (\n        load_data()\n        >> (analyze_a() & analyze_b())\n        >> publish()\n    )",
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
        )
    )
    return


@app.cell(hide_code=True)
def _(WORKFLOW_NODE_TYPES_DAG_PATH, mo):
    mo.image(
        WORKFLOW_NODE_TYPES_DAG_PATH,
        alt="Avalanche workflow DAG annotated with node types",
        width="100%",
        rounded=True,
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.vstack(
        (
            mo.md("## 6. Implement the remaining steps"),
            mo.md(
                "The source exposes the bundled workbook. A second agent independently "
                "detects risks, deterministic Python composes the approved review, two "
                "capability-specific agents render its artifacts, and the destination "
                "publishes the files together."
            ),
        )
    )
    return


@app.cell(hide_code=True)
def _(BaseModel, Field, Literal):
    CustomerSegment = Literal["enterprise", "mid_market", "small_business", "startup"]
    RiskSeverity = Literal["low", "medium", "high"]

    class EvidenceReference(BaseModel):
        feedback_id: str
        account_id: str
        sheet: str
        row_number: int = Field(ge=2)
        excerpt: str

    class Theme(BaseModel):
        name: str
        summary: str
        feedback_count: int = Field(ge=2)
        account_count: int = Field(ge=2)
        segment_counts: dict[CustomerSegment, int]
        evidence: list[EvidenceReference] = Field(min_length=2)

    class ThemeReport(BaseModel):
        feedback_rows_analyzed: int = Field(ge=1)
        themes: list[Theme]

    class AccountExposure(BaseModel):
        account_id: str
        account_name: str
        customer_segment: CustomerSegment
        arr_usd: int = Field(ge=0)
        renewal_date: str = Field(pattern=r"^\d{4}-\d{2}-\d{2}$")

    class Risk(BaseModel):
        issue: str
        severity: RiskSeverity
        rationale: str
        product_areas: list[str] = Field(min_length=1)
        affected_accounts: list[AccountExposure] = Field(min_length=1)
        arr_at_risk_usd: int = Field(ge=0)
        evidence: list[EvidenceReference] = Field(min_length=1)

    class RiskReport(BaseModel):
        feedback_rows_analyzed: int = Field(ge=1)
        risks: list[Risk]

    class ProductReview(BaseModel):
        feedback_rows_analyzed: int = Field(ge=1)
        themes: list[Theme]
        risks: list[Risk]

    class PublishedReviewPack(BaseModel):
        workbook_path: str
        brief_path: str

    return ProductReview, PublishedReviewPack, RiskReport, ThemeReport


@app.cell(hide_code=True)
def _(
    BRIEF_OUTPUT_DIR,
    FEEDBACK_WORKBOOK_PATH,
    File,
    MODEL,
    ProductReview,
    PublishedReviewPack,
    RiskReport,
    SUB_MODEL,
    ThemeReport,
    WORKBOOK_OUTPUT_DIR,
    ava,
    evidence_coding_skill,
):
    @ava.source
    def load_feedback_workbook() -> File:
        return File(path=str(FEEDBACK_WORKBOOK_PATH))

    class ExtractThemes(ava.Signature):
        """Find recurring cross-account product needs in the feedback workbook.

        Inspect and validate the workbook before analysis. Use focused predict() calls for
        semantic coding and Python for grouping, deduplication, and counts. Require every
        theme to span at least two accounts and preserve workbook evidence.
        """

        workbook: File = ava.InputField(desc="Complete feedback workbook.")
        report: ThemeReport = ava.OutputField(desc="Recurring evidence-backed themes.")

    @ava.agent_step(
        ExtractThemes,
        lm=MODEL,
        sub_lm=SUB_MODEL,
        skills=[ava.agent.skills.spreadsheet, evidence_coding_skill],
    )
    async def extract_themes(workbook: File, *, agent: ava.Agent) -> ThemeReport:
        return ThemeReport.model_validate((await agent(workbook=workbook)).report)

    class DetectRisks(ava.Signature):
        """Identify material account risks grounded in the feedback workbook."""

        workbook: File = ava.InputField(desc="Complete feedback workbook.")
        report: RiskReport = ava.OutputField(desc="Evidence-backed account risks.")

    @ava.agent_step(
        DetectRisks,
        lm=MODEL,
        sub_lm=SUB_MODEL,
        skills=[ava.agent.skills.spreadsheet, evidence_coding_skill],
    )
    async def detect_risks(workbook: File, *, agent: ava.Agent) -> RiskReport:
        return RiskReport.model_validate((await agent(workbook=workbook)).report)

    @ava.step
    def compose_product_review(
        themes: ThemeReport,
        risks: RiskReport,
    ) -> ProductReview:
        if themes.feedback_rows_analyzed != risks.feedback_rows_analyzed:
            raise ValueError("theme and risk reports analyzed different row counts")
        return ProductReview(
            feedback_rows_analyzed=themes.feedback_rows_analyzed,
            themes=themes.themes,
            risks=risks.risks,
        )

    class BuildReviewWorkbook(ava.Signature):
        """Render the approved review as a validated Excel workbook."""

        source_workbook: File = ava.InputField(desc="Original evidence workbook.")
        review: ProductReview = ava.InputField(desc="Approved review to render.")
        workbook: File = ava.OutputField(desc="Rendered product_review.xlsx.")

    @ava.agent_step(
        BuildReviewWorkbook,
        lm=MODEL,
        sub_lm=SUB_MODEL,
        skills=[ava.agent.skills.spreadsheet],
        output_dir=WORKBOOK_OUTPUT_DIR,
    )
    async def build_review_workbook(
        source_workbook: File,
        review: ProductReview,
        *,
        agent: ava.Agent,
    ) -> File:
        return File.model_validate(
            (await agent(source_workbook=source_workbook, review=review)).workbook
        )

    class WriteExecutiveBrief(ava.Signature):
        """Render the approved review as a concise Word executive brief."""

        review: ProductReview = ava.InputField(desc="Approved review to render.")
        brief: File = ava.OutputField(desc="Rendered executive_brief.docx.")

    @ava.agent_step(
        WriteExecutiveBrief,
        lm=MODEL,
        sub_lm=SUB_MODEL,
        skills=[ava.agent.skills.docx],
        output_dir=BRIEF_OUTPUT_DIR,
    )
    async def write_executive_brief(
        review: ProductReview,
        *,
        agent: ava.Agent,
    ) -> File:
        return File.model_validate((await agent(review=review)).brief)

    @ava.dest
    def publish_review_pack(workbook: File, brief: File) -> PublishedReviewPack:
        if workbook.path is None or brief.path is None:
            raise ValueError("review artifacts must have host file paths")
        return PublishedReviewPack(
            workbook_path=workbook.path,
            brief_path=brief.path,
        )

    return (
        ExtractThemes,
        build_review_workbook,
        compose_product_review,
        detect_risks,
        extract_themes,
        load_feedback_workbook,
        publish_review_pack,
        write_executive_brief,
    )


@app.cell(hide_code=True)
def _(mo, source_code):
    mo.vstack(
        (
            mo.md("### Load the workbook"),
            mo.ui.code_editor(
                value=source_code("flow.py", "load_feedback_workbook"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
            mo.md("### Detect risks independently"),
            mo.ui.code_editor(
                value=source_code("signature.py", "DetectRisks")
                + "\n\n\n"
                + source_code("flow.py", "detect_risks"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
            mo.md("### Join the analyses deterministically"),
            mo.ui.code_editor(
                value=source_code("flow.py", "compose_product_review"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
            mo.md("### Build the review workbook"),
            mo.ui.code_editor(
                value=source_code("signature.py", "BuildReviewWorkbook")
                + "\n\n\n"
                + source_code("flow.py", "build_review_workbook"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
            mo.md("### Write the executive brief"),
            mo.ui.code_editor(
                value=source_code("signature.py", "WriteExecutiveBrief")
                + "\n\n\n"
                + source_code("flow.py", "write_executive_brief"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
            mo.md("### Publish the review pack"),
            mo.ui.code_editor(
                value=source_code("flow.py", "publish_review_pack"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
        )
    )
    return


@app.cell(hide_code=True)
def _(
    ava,
    build_review_workbook,
    compose_product_review,
    detect_risks,
    extract_themes,
    load_feedback_workbook,
    publish_review_pack,
    write_executive_brief,
):
    @ava.workflow
    def feedback_review():
        return (
            (workbook := load_feedback_workbook())
            >> ((themes := extract_themes(workbook)) & (risks := detect_risks(workbook)))
            >> (review := compose_product_review(themes, risks))
            >> (build_review_workbook(workbook, review) & write_executive_brief(review))
            >> publish_review_pack()
        )

    return


@app.cell(hide_code=True)
def _(mo, source_code):
    mo.vstack(
        (
            mo.md("## 7. Declare the workflow"),
            mo.md(
                "The nodes already own their individual responsibilities. The workflow "
                "body only declares their dependencies: `>>` orders stages, while `&` "
                "forms the independent analysis and rendering branches."
            ),
            mo.ui.code_editor(
                value=source_code("flow.py", "feedback_review"),
                language="python",
                disabled=True,
                show_copy_button=True,
            ),
        )
    )
    return


if __name__ == "__main__":
    app.run()
